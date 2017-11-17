import smtplib
from Weather import Weather
import openpyxl
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText


class cue:
    """ Actual cuer and takes in a list
        of desired items in email/
    """
    contact = dict(Jason='cue.me.today@gmail.com')

    def __init__(self, name, feats):
        self.mailer = smtplib.SMTP('smtp.gmail.com', 587)
        self.mailer.ehlo()
        self.mailer.starttls()
        self.mailer.login('cue.me.today@gmail.com', 'okay12345')

        self.name = name
        self.subject = "Your upcoming day, %s!" % name
        self.msg = MIMEMultipart('alternative')
        self.msg['Subject'] = "Your upcoming day, %s!" % name
        self.msg['From'] = self.msg['To'] = 'cue.me.today@gmail.com'

        # self.msg = "Dear %s," % name
        self.feats = feats

        self.composer()

    def composer(self):
        """ Creates the string used
            to send the email.
        """

        text = "Hi!\nHow are you?\nHere is the link you wanted:\nhttps://www.python.org"

        html = """\
        <html>
          <head>
          <link href="https://fonts.googleapis.com/css?family=Source+Sans+Pro" rel="stylesheet">
          </head>
          <style>

          h1 {
            font-family: 'Source Sans Pro', sans-serif;
          }

          p {
            font-family: 'Source Sans Pro', sans-serif;
          }


          </style>
            <body style="margin: 0; padding: 0;">
                <table align="center" border="1" bordercolor=BLACK cellpadding="50" cellspacing="0" width="600">
                 <tr>
                  <td bgcolor="#70bbd9">
                   <h1>Good Morning, Jason!</h1>
                   <p>Thursday, November 16, 2017.<br> Here's what's coming up today.</p>
                  </td>
                 </tr>
                 <tr>
                   <td bgcolor="#ffffff" style="padding: 40px 30px 40px 30px;">
                    <table border="1" cellpadding="0" cellspacing="0" width="100%%">
                     <tr>
                      <td>
                       %s
                      </td>
                     </tr>
                     <tr>
                      <td>
                       %s
                       <iframe src="https://calendar.google.com/calendar/embed?src=cue.me.today%%40gmail.com&ctz=America%%2FLos_Angeles" style="border: 0" width="800" height="600" frameborder="0" scrolling="no"></iframe>             </tr>
                      </td>
                     </tr>
                     <tr>
                      <td>
                        <table border="1" cellpadding="0" cellspacing="0" width="100%%">
                         <tr>
                          <td width="50%%" valign="top">
                           some sort of interaction
                          </td>
                          <td style="font-size: 0; line-height: 0;" width="0%%">
                           &nbsp;
                          </td>
                          <td width="260" valign="top">
                            some sort of interaction2
                          </td>
                         </tr>
                        </table>
                      </td>
                     </tr>
                    </table>
                   </td>
                 </tr>
                 <tr>
                  <td bgcolor="#ee4c50">
                   <p>Have a good day and remember to bring your umbrella!</p>
                  </td>
                 </tr>
                </table>
            </body>
        </html>
        """ % (Weather(self.name).compose(), reminders(self.name).compose())
        print(html)

        part1 = MIMEText(text, 'plain')
        part2 = MIMEText(html, 'html')
        self.msg.attach(part1)
        self.msg.attach(part2)

        """
        if 'weather' in self.feats :
            self.msg += weather(self.name).compose()
        if 'task' in self.feats:
            self.msg += task(self.name).compose()
        """

    def mail(self):
        """ Sends the daily email notifiation
            with wanted features.
        """
        """
        self.mailer.sendmail('cue.me.today@gmail.com',
                             self.contact[self.name],
                             'Subject: %s<br>%s' % (self.subject, self.msg))        
        """
        print('sent')
        self.mailer.sendmail('cue.me.today@gmail.com',
                             self.contact[self.name],
                             self.msg.as_string())


class reminders:  # should move to abstract
    """ Reads from excel of tasks.
    """

    def __init__(self, name):
        self.wb = openpyxl.load_workbook('%s.xlsx' % name)
        self.sheet = self.wb.get_sheet_by_name('Sheet1')
        self.compose()

    def compose(self):
        msg = ''
        i = 2
        task = self.sheet.cell(row=i, column=1).value
        while task != None:  # due date should be adjusted in this
            due = self.sheet.cell(row=i, column=2).value
            msg += "%s due on %s<br>" % (task, due)
            i += 1
            task = self.sheet.cell(row=i, column=1).value
        msg = "Today you have %d things to do:<br>" % (i - 2) + msg
        return msg


class weather:  # implemented class
    """ Gives weather.
    """

    def __init__(self, name):
        owm = pyowm.OWM('c56ba86f15e44e73a6e50e64c99b06df')
        self.fc = owm.three_hours_forecast("Berkeley,us")
        f = self.fc.get_forecast();
        self.lst = f.get_weathers()[0:6]

        # fixme this is doing for all five days
        self.hot = self.fc.most_hot().get_temperature('fahrenheit')['temp']
        self.ht = self.fc.most_hot().get_reference_time(timeformat='iso')

        self.cold = self.fc.most_cold().get_temperature('fahrenheit')['temp']
        self.ct = self.fc.most_cold().get_reference_time(timeformat='iso')

        self.compose()

    def compose(self):  # fixme --> includes 5 days
        msg = "Today, the high will be %sF around " \
              "%s and the low will be %sF around %s.<br>" \
              % (self.hot, self.ht, self.cold, self.ct)
        if self.fc.will_have_rain():
            r = self.fc.when_rain()[0];
            msg += "It will also rain today around %s. Bring an umbrella!<br>" % r.get_reference_time(timeformat='iso')
        return msg


def main():
    """ Read the correct profile.
    """
    q = cue('Jason', ['task'])
    q.mail();


main()
