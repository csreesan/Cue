import openpyxl
from email.mime.application import MIMEApplication


class Reminders:
    """ Reads from excel of tasks.
    """

    def __init__(self, sheet_name):
        self.sheet_name = sheet_name
        self.wb = openpyxl.load_workbook('%s.xlsx' % self.sheet_name)
        self.sheet = self.wb.get_sheet_by_name('Sheet1')
        self.attachments = [];

    def compose(self):
        msg = ''
        i = 2
        task = self.sheet.cell(row=i, column=1).value
        while task:  # due date should be adjusted in this
            due = self.sheet.cell(row=i, column=2).value
            msg += "%s due on %s<br>" % (task, due)
            if self.sheet.cell(row=i, column=3).value != None:
                cell = self.sheet.cell(row=i, column=3).value
                msg += "Attached is %s." % cell
                self.attach(cell)
            i += 1
            task = self.sheet.cell(row=i, column=1).value
        msg = "Today you have %d things to do:<br>" % (i - 2) + msg
        return msg

    def attach(self, cell):
        fp = open(cell, 'rb')
        file = MIMEApplication(fp.read(), _subtype="pdf")
        fp.close()
        file.add_header('Content-Disposition', 'attachment', filename=cell)
        self.attachments.append(file)

    def get_attachments(self):
        return self.attachments;
